from wxpy import*
import pygame
from openpyxl import load_workbook
import requests
import datetime
import time
from datetime import datetime, timedelta
import win32com.client as win
import win32com.client as wc
import win32api
import win32con
import json
import urllib.request
from urllib.parse import urlencode

speak = win.Dispatch("SAPI.SpVoice")
bot = Bot()
girl_friend=bot.search('YQD')[0]
self_friend=bot.search('XHR')[0]

global zhuangtai
zhuangtai = '尚未设置状态信息'

@bot.register()
def recv_send_msg(recv_msg):
    global zhuangtai
    print(recv_msg) # recv_msg.text取得文本
    if recv_msg.sender == girl_friend:
        msg = str(recv_msg)
        if "包子" in msg:
            girl_friend.send('你好啊')
        elif "骗" in msg:
            girl_friend.send('包子没有骗小乙哦，可能是我又出Bug了')
            speak.Speak('小呆，包子又出Bug了，记得做好记录')
        elif "小呆在干嘛" in msg:
            girl_friend.send('小呆在'+zhuangtai)
            speak.Speak('小呆，小乙有点想你，在问你在干嘛')
        elif "懒猪" in msg:
            girl_friend.send('正在播放音乐')
            import pygame
            pygame.mixer.init()
            print("正在播放音乐")
            track = pygame.mixer.music.load(r'D:/Loveless.mp3')
            pygame.mixer.music.play()
            time.sleep(60)
            pygame.mixer.music.stop()
            girl_friend.send('播放结束')
            print("播放结束")
        elif "叫醒" in msg:
            girl_friend.send('正在播放音乐')
            import pygame
            pygame.mixer.init()
            print("正在播放音乐")
            track = pygame.mixer.music.load(r'D:/Loveless.mp3')
            pygame.mixer.music.play()
            time.sleep(60)
            pygame.mixer.music.stop()
            girl_friend.send('播放结束')
            print("播放结束")
        elif "叫小呆起床" in msg:
            girl_friend.send('正在播放音乐')
            import pygame
            pygame.mixer.init()
            print("正在播放音乐")
            track = pygame.mixer.music.load(r'D:/Loveless.mp3')
            pygame.mixer.music.play()
            time.sleep(60)
            pygame.mixer.music.stop()
            girl_friend.send('播放结束')
            print("播放结束")
        elif "睡觉" in msg:
            speak.Speak('小乙说时间比较晚了该睡觉了')
            girl_friend.send('提醒发送成功')
        elif "小呆在干嘛" in msg:
            girl_friend.send('小呆在'+zhuangtai)
            speak.Speak('小呆，小乙有点想你，在问你在干嘛')
        elif "早上好" in msg:
            pre = datetime(2019, 4, 13)  # 年月日时分秒 微秒
            now = datetime.now()
            dst = now - pre
            dst = str(dst)
            dst = dst[0:3]
            msg0 = '早上好呀，你们已经相爱' + dst + '天了哦。小呆在' + zhuangtai + '。'

            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州今日' + info1['forecasts'][0]['casts'][0][
                'dayweather'] + '，气温:' + ' ' + info1['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][0]['daytemp'] + '度。'
            msg2 = '阜宁今日' + info2['forecasts'][0]['casts'][0][
                'dayweather'] + '，气温:' + ' ' + info2['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][0]['daytemp'] + '度。'
            print(msg0)
            print(msg1)
            print(msg2)
            girl_friend.send(msg0 + msg2 + msg1 + '新的一天要元气满满哦。')
        elif "今天天气" in msg:
            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州今日天气:' + ' ' + info1['forecasts'][0]['casts'][0][
                'dayweather'] + ' ' + ':气温:' + ' ' + info1['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][0]['daytemp'] + '度'
            msg2 = '阜宁今日天气:' + ' ' + info2['forecasts'][0]['casts'][0][
                'dayweather'] + ' ' + ':气温:' + ' ' + info2['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][0]['daytemp'] + '度'
            print(msg1)
            print(msg2)
            girl_friend.send(msg1)
            girl_friend.send(msg2)
        elif "明天天气" in msg:
            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州' + '明日天气:' + ' ' + info1['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温:' + ' ' + info1['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][1]['daytemp'] + '度'
            msg2 = '阜宁' + '明日天气:' + ' ' + info2['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温' + ' ' + info2['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][1]['daytemp'] + '度'
            print(msg1)
            print(msg2)
            girl_friend.send(msg1)
            girl_friend.send(msg2)
        elif "语音播报" in msg:
            msg1 = msg[0:3]
            msg2 = msg[6:10]
            msg3 = msg[11:-6]
            #print(msg1)  # 截取用户名
            #print(msg2)  # 截取消息类型
            #print(msg3)  # 截取消息内容
            speak.speak('小乙发来语音消息：'+msg3)
            girl_friend.send('语音消息已成功播报')
        elif "查菜谱" in msg:
            daicha = msg[10:-6]
            appkey = "************************"   ###聚合数据
            url = "http://apis.juhe.cn/cook/query.php"
            params = {
                "menu": daicha,  # 需要查询的菜谱名
                "key": appkey,  # 应用APPKEY(应用详细页查询)
                "dtype": "",  # 返回数据的格式,xml或json，默认json
                "pn": "",  # 数据返回起始下标
                "rn": "1",  # 数据返回条数，最大30
                "albums": "",  # albums字段类型，1字符串，默认数组
            }
            params = urlencode(params)
            f = urllib.request.urlopen("%s?%s" % (url, params))

            content = f.read()
            res = json.loads(content)
            if res:
                error_code = res["error_code"]
                #输出数据
                if error_code == 0:
                    # 成功请求
                    self_friend.send(daicha+'做法：')
                    try:
                        print(res["result"]["data"][0]["steps"][0]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][0]["step"])
                    except:
                        print('步骤1：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][1]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][1]["step"])
                    except:
                        print('步骤2：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][2]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][2]["step"])
                    except:
                        print('步骤3：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][3]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][3]["step"])
                    except:
                        print('步骤4：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][4]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][4]["step"])
                    except:
                        print('步骤5：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][5]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][5]["step"])
                    except:
                        print('步骤6：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][6]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][6]["step"])
                    except:
                        print('步骤7：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][7]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][7]["step"])
                    except:
                        print('步骤8：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][8]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][8]["step"])
                    except:
                        print('步骤9：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][9]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][9]["step"])
                    except:
                        print('步骤10：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][10]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][10]["step"])
                    except:
                        print('步骤11：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][11]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][11]["step"])
                    except:
                        print('步骤12：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][12]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][12]["step"])
                    except:
                        print('步骤13：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][13]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][13]["step"])
                    except:
                        print('步骤14：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][14]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][14]["step"])
                    except:
                        print('步骤15：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][15]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][15]["step"])
                    except:
                        print('步骤16：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][16]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][16]["step"])
                    except:
                        print('步骤17：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][17]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][17]["step"])
                    except:
                        print('步骤18：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][18]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][18]["step"])
                    except:
                        print('步骤19：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][19]["step"])
                        girl_friend.send(res["result"]["data"][0]["steps"][19]["step"])
                    except:
                        print('步骤20：步骤结束')
                else:
                    print("%s:%s" % (res["error_code"], res["reason"]))
                    girl_friend.send('暂未找到相匹配菜谱数据，可更换关键词或转用浏览器查询')
            else:
                print("request api error")
                girl_friend.send('包子功能故障，已联系小呆修理啦')
                self_friend.send('包子菜谱功能故障啦，请查看运行日志进行修正哦！')
        elif "记日记" in msg:
            datatime = str(time.strftime("%Y%m%d %H:%M:%S", time.localtime()))
            wb = load_workbook('D:\\riji.xlsx')  # 创建/读取文件对象
            ws = wb.active  # 获取第一个sheet
            # 字段截取
            msg1 = msg[0:3]
            msg2 = msg[6:9]
            msg3 = msg[10:-6]

            # print(msg1)  # 截取用户名
            # print(msg2)  # 截取日程类型
            # print(msg3)  # 截取日程内容

            ws.append([msg1, msg2, msg3, datatime])  # 写入多个单元格
            # Save the file
            wb.save('D:\\riji.xlsx')

            speak.Speak('小乙已建立新日记')
            girl_friend.send('新日记已建立'+msg3)
            print('新日记已建立'+msg3)
        else:
            msg = '包子还没有学会哦，已经通知小呆教我啦，可先尝试其他指令哦'
            girl_friend.send(msg)
    elif recv_msg.sender == self_friend:
        msg = str(recv_msg)
        if "当前状态" in msg:
            zhuangtai = msg[10:-6]
            speak.Speak('当前状态已设置为:' + zhuangtai)
        elif "叫小呆起床" in msg:
            self_friend.send('正在播放音乐')
            import pygame
            pygame.mixer.init()
            print("正在播放音乐")
            track = pygame.mixer.music.load(r'D:/Loveless.mp3')
            
            pygame.mixer.music.play()
            time.sleep(60)
            pygame.mixer.music.stop()
            self_friend.send('播放结束')
            print("播放结束")
        elif "早上好" in msg:
            pre = datetime(2019, 4, 13)  # 年月日时分秒 微秒
            now = datetime.now()
            dst = now - pre
            dst = str(dst)
            dst = dst[0:3]
            msg0 = '早上好，今天是你们已经相爱的第' + dst + '天。'

            pre = datetime(2019, 6, 28)  # 年月日时分秒 微秒
            now = datetime.now()
            dst = now - pre
            dst = str(dst)
            dst = dst[0:2]
            msg4 = '今天是50天学习计划第' + dst + '天。'


            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州今日' + info1['forecasts'][0]['casts'][0][
                'dayweather'] + '，气温:' + ' ' + info1['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][0]['daytemp'] + '度。'
            msg2 = '阜宁今日' + info2['forecasts'][0]['casts'][0][
                'dayweather'] + '，气温:' + ' ' + info2['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][0]['daytemp'] + '度。'
            if "雨" in info1['forecasts'][0]['casts'][0]['dayweather']:
                msg5 = '今日出门记得带伞'
            else:
                msg5 = '.'
            if "雨" in info2['forecasts'][0]['casts'][0]['dayweather']:
                msg6 = '记得提醒小乙出门带伞'
            else:
                msg6 = '.'

            print(msg0)
            print(msg1)
            print(msg2)
            self_friend.send(msg0+msg4+msg2+msg1+msg5+msg6)
            speak.Speak(msg0+msg4+msg2+msg1+msg5+msg6)

            excel_app = wc.Dispatch('Excel.Application')
            workbook = excel_app.Workbooks.Open(r'D:\\richengbiao.xlsx')
            deadtime = str(time.strftime("%m%d", time.localtime()))
            i = 2
            while True:
                code1 = workbook.Worksheets('Sheet1').Cells(i, 4).Value
                if code1 == None:
                    break
                code1 = int(code1)
                #print(code1)
                if code1 == int(deadtime):
                    print(workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                    self_friend.send('记得'+workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                    speak.Speak('记得'+workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                # print(i)
                i += 1
            excel_app.Quit()
        elif "晚上好" in msg:
            pre = datetime(2019, 4, 13)  # 年月日时分秒 微秒
            now = datetime.now()
            dst = now - pre
            dst = str(dst)
            dst = dst[0:2]
            msg0 = '晚上好，你们已经相爱了' + dst + '天。'

            pre = datetime(2019, 6, 28)  # 年月日时分秒 微秒
            now = datetime.now()
            dst = now - pre
            dst = str(dst)
            dst = dst[0:2]
            msg4 = '50天学习计划已经进行了' + dst + '天。'

            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州' + '明日天气:' + ' ' + info1['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温:' + ' ' + info1['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][1]['daytemp'] + '度'
            msg2 = '阜宁' + '明日天气:' + ' ' + info2['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温' + ' ' + info2['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][1]['daytemp'] + '度'
            print(msg1)
            print(msg2)
            if "雨" in info1['forecasts'][0]['casts'][1]['dayweather']:
                msg5 = '明天出门记得带伞'
            else:
                msg5 = '.'
            if "雨" in info2['forecasts'][0]['casts'][1]['dayweather']:
                msg6 = '记得提醒小乙明天出门带伞'
            else:
                msg6 = '.'
            self_friend.send(msg0+msg4+msg2+msg1+msg5+msg6)
            speak.Speak(msg0+msg4+msg2+msg1+msg5+msg6)

            excel_app = wc.Dispatch('Excel.Application')
            workbook = excel_app.Workbooks.Open(r'D:\\richengbiao.xlsx')
            deadtime = str(time.strftime("%m%d", time.localtime(time.time()+86400)))
            i = 2
            while True:
                code1 = workbook.Worksheets('Sheet1').Cells(i, 4).Value
                if code1 == None:
                    break
                code1 = int(code1)
                # print(code1)
                if code1 == int(deadtime):
                    print(workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                    self_friend.send('记得明天' + workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                    speak.Speak('记得明天' + workbook.Worksheets('Sheet1').Cells(i, 3).Value)
                # print(i)
                i += 1
            excel_app.Quit()
        elif "今天天气" in msg:
            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州今日天气:' + ' ' + info1['forecasts'][0]['casts'][0][
                'dayweather'] + ' ' + ':气温:' + ' ' + info1['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][0]['daytemp'] + '度'
            msg2 = '阜宁今日天气:' + ' ' + info2['forecasts'][0]['casts'][0][
                'dayweather'] + ' ' + ':气温:' + ' ' + info2['forecasts'][0]['casts'][0]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][0]['daytemp'] + '度'
            print(msg1)
            print(msg2)
            self_friend.send(msg1)
            self_friend.send(msg2)
            speak.Speak(msg1)
            speak.Speak(msg2)
        elif "明天天气" in msg:
            req1 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320505&extensions=all")
            req2 = requests.get(
                "http://restapi.amap.com/v3/weather/weatherInfo?key=************************&city=320923&extensions=all")
            info1 = dict(req1.json())
            info2 = dict(req2.json())
            msg1 = '苏州' + '明日天气:' + ' ' + info1['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温:' + ' ' + info1['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info1['forecasts'][0]['casts'][1]['daytemp'] + '度'
            msg2 = '阜宁' + '明日天气:' + ' ' + info2['forecasts'][0]['casts'][1][
                'dayweather'] + ' ' + ':气温' + ' ' + info2['forecasts'][0]['casts'][1]['nighttemp'] + '~' + \
                   info2['forecasts'][0]['casts'][1]['daytemp'] + '度'
            print(msg1)
            print(msg2)
            self_friend.send(msg1)
            self_friend.send(msg2)
            speak.Speak(msg1)
            speak.Speak(msg2)
        elif "提醒我" in msg:
            deadtime = str(time.strftime("%Y%m%d %H:%M:%S", time.localtime()))
            wb = load_workbook('D:\\richengbiao.xlsx')  # 创建/读取文件对象
            ws = wb.active  # 获取第一个sheet
            # 字段截取
            msg1 = msg[0:3]
            msg2 = msg[6:9]
            msg3 = msg[10:-12]
            msg4 = msg[-11:-6]

            # print(msg1)  # 截取用户名
            # print(msg2)  # 截取日程类型
            # print(msg3)  # 截取日程内容
            # print(msg4)  # 截取提醒时间

            ws.append([msg1, msg2, msg3, msg4, deadtime])  # 写入多个单元格
            # Save the file
            wb.save('D:\\richengbiao.xlsx')

            speak.Speak('新日程已建立，将会于' + msg4 + '日提醒你' + msg3)
            self_friend.send('新日程已建立，将会于' + msg4 + '日提醒你' + msg3)
            print('新日程已建立，将会于' + msg4 + '日提醒你' + msg3)
        elif "语音播报" in msg:
            msg1 = msg[0:3]
            msg2 = msg[6:10]
            msg3 = msg[11:-6]
            #print(msg1)  # 截取用户名
            #print(msg2)  # 截取消息类型
            #print(msg3)  # 截取消息内容
            speak.speak(msg3)
        elif "升高音量" in msg:
            from ctypes import windll
            WM_APPCOMMAND = 0x319
            APPCOMMAND_VOLUME_UP = 0x0a
            APPCOMMAND_VOLUME_DOWN = 0x09
            APPCOMMAND_VOLUME_MUTE = 0x08
            hwnd = windll.user32.GetForegroundWindow()
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
        elif "增大音量" in msg:
            from ctypes import windll
            WM_APPCOMMAND = 0x319
            APPCOMMAND_VOLUME_UP = 0x0a
            APPCOMMAND_VOLUME_DOWN = 0x09
            APPCOMMAND_VOLUME_MUTE = 0x08
            hwnd = windll.user32.GetForegroundWindow()
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_UP * 0x10000)
        elif "降低音量" in msg:
            from ctypes import windll
            WM_APPCOMMAND = 0x319
            APPCOMMAND_VOLUME_UP = 0x0a
            APPCOMMAND_VOLUME_DOWN = 0x09
            APPCOMMAND_VOLUME_MUTE = 0x08
            hwnd = windll.user32.GetForegroundWindow()
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
        elif "减小音量" in msg:
            from ctypes import windll
            WM_APPCOMMAND = 0x319
            APPCOMMAND_VOLUME_UP = 0x0a
            APPCOMMAND_VOLUME_DOWN = 0x09
            APPCOMMAND_VOLUME_MUTE = 0x08
            hwnd = windll.user32.GetForegroundWindow()
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_DOWN * 0x10000)
        elif "静音" in msg:
            from ctypes import windll
            WM_APPCOMMAND = 0x319
            APPCOMMAND_VOLUME_UP = 0x0a
            APPCOMMAND_VOLUME_DOWN = 0x09
            APPCOMMAND_VOLUME_MUTE = 0x08
            hwnd = windll.user32.GetForegroundWindow()
            windll.user32.PostMessageA(hwnd, WM_APPCOMMAND, 0, APPCOMMAND_VOLUME_MUTE * 0x10000)
        elif "下一首" in msg:
            win32api.keybd_event(0xB0, 0, 0, 0)  # ‘下一首’位码是 0xB0
            win32api.keybd_event(0xB0, 0, win32con.KEYEVENTF_KEYUP, 0)
        elif "上一首" in msg:
            win32api.keybd_event(0xB1, 0, 0, 0)  # ‘下一首’位码是 0xB1
            win32api.keybd_event(0xB1, 0, win32con.KEYEVENTF_KEYUP, 0)
        elif "暂停" in msg:
            win32api.keybd_event(0xB3, 0, 0, 0)  # ‘下一首’位码是 0xB3
            win32api.keybd_event(0xB3, 0, win32con.KEYEVENTF_KEYUP, 0)
        elif "播放" in msg:
            win32api.keybd_event(0xB3, 0, 0, 0)  # ‘下一首’位码是 0xB3
            win32api.keybd_event(0xB3, 0, win32con.KEYEVENTF_KEYUP, 0)
        elif "查菜谱" in msg:
            daicha = msg[10:-6]
            appkey = "************************"   ###聚合数据
            url = "http://apis.juhe.cn/cook/query.php"
            params = {
                "menu": daicha,  # 需要查询的菜谱名
                "key": appkey,  # 应用APPKEY(应用详细页查询)
                "dtype": "",  # 返回数据的格式,xml或json，默认json
                "pn": "",  # 数据返回起始下标
                "rn": "1",  # 数据返回条数，最大30
                "albums": "",  # albums字段类型，1字符串，默认数组
            }
            params = urlencode(params)

            f = urllib.request.urlopen("%s?%s" % (url, params))

            content = f.read()
            res = json.loads(content)
            if res:
                error_code = res["error_code"]
                #输出数据
                if error_code == 0:
                    # 成功请求
                    self_friend.send(daicha+'做法：')
                    try:
                        print(res["result"]["data"][0]["steps"][0]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][0]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][0]["step"])
                    except:
                        print('步骤1：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][1]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][1]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][1]["step"])
                    except:
                        print('步骤2：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][2]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][2]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][2]["step"])
                    except:
                        print('步骤3：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][3]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][3]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][3]["step"])
                    except:
                        print('步骤4：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][4]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][4]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][4]["step"])
                    except:
                        print('步骤5：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][5]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][5]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][5]["step"])
                    except:
                        print('步骤6：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][6]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][6]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][6]["step"])
                    except:
                        print('步骤7：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][7]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][7]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][7]["step"])
                    except:
                        print('步骤8：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][8]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][8]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][8]["step"])
                    except:
                        print('步骤9：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][9]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][9]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][9]["step"])
                    except:
                        print('步骤10：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][10]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][10]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][10]["step"])
                    except:
                        print('步骤11：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][11]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][11]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][11]["step"])
                    except:
                        print('步骤12：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][12]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][12]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][12]["step"])
                    except:
                        print('步骤13：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][13]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][13]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][13]["step"])
                    except:
                        print('步骤14：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][14]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][14]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][14]["step"])
                    except:
                        print('步骤15：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][15]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][15]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][15]["step"])
                    except:
                        print('步骤16：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][16]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][16]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][16]["step"])
                    except:
                        print('步骤17：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][17]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][17]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][17]["step"])
                    except:
                        print('步骤18：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][18]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][18]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][18]["step"])
                    except:
                        print('步骤19：步骤结束')
                    try:
                        print(res["result"]["data"][0]["steps"][19]["step"])
                        self_friend.send(res["result"]["data"][0]["steps"][19]["step"])
                        speak.speak(res["result"]["data"][0]["steps"][19]["step"])
                    except:
                        print('步骤20：步骤结束')
                else:
                    print("%s:%s" % (res["error_code"], res["reason"]))
                    self_friend.send('暂未找到相匹配菜谱数据，将询问蒸饺')
                    speak.speak('暂未找到相匹配菜谱数据')
                    speak.speak('蒸饺蒸饺：')
                    time.sleep(2)
                    speak.speak(daicha + '怎么做？')
            else:
                print("request api error")
        elif "记日记" in msg:
            datatime = str(time.strftime("%Y%m%d %H:%M:%S", time.localtime()))
            wb = load_workbook('D:\\riji.xlsx')  # 创建/读取文件对象
            ws = wb.active  # 获取第一个sheet
            # 字段截取
            msg1 = msg[0:3]
            msg2 = msg[6:9]
            msg3 = msg[10:-6]

            # print(msg1)  # 截取用户名
            # print(msg2)  # 截取日程类型
            # print(msg3)  # 截取日程内容

            ws.append([msg1, msg2, msg3, datatime])  # 写入多个单元格
            # Save the file
            wb.save('D:\\riji.xlsx')

            speak.Speak('新日记已建立')
            self_friend.send('新日记已建立'+msg3)
            print('新日记已建立'+msg3)
        elif "查看所有日记" in msg:
            excel_app = wc.Dispatch('Excel.Application')
            workbook = excel_app.Workbooks.Open(r'D:\\riji.xlsx')

            i = 1
            while True:
                name1 = workbook.Worksheets('Sheet1').Cells(i, 1).Value
                text1 = workbook.Worksheets('Sheet1').Cells(i, 3).Value
                data1 = workbook.Worksheets('Sheet1').Cells(i, 4).Value
                if name1 == None:
                    break

                rijineirong = name1+data1+'\n'+text1

                print(rijineirong)
                self_friend.send(rijineirong)

                # print(i)
                i += 1
            excel_app.Quit()
        else:
            msg = '包子还没有学会哦，可尝试其他指令'
            self_friend.send(msg)
    else:
        self_friend.send('  ')
            
embed()









